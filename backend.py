import math
import random
import openpyxl
import requests
import time
from typing import List, Dict
import pandas as pd
from datetime import datetime, timedelta
import numpy as np
import calculate_buffet_score
import calculate_multiplier_score
import calculate_vwap_score
import calculate_dividends_score
import calculate_expected_returns_and_volatility
from bs4 import BeautifulSoup


class Stock:
    def __init__(self, ticker: str):
        self.ticker = ticker
        self.sector = self.get_sector_name()         # Определить сектор через API
        self.get_moex_data()
        
        # Методы расчета
        expected_returns_and_volatility = calculate_expected_returns_and_volatility.get_expected_returns_and_volatility(self.ticker)
        self.return_5y = expected_returns_and_volatility[0]  # Ожидаемая доходность за 5 лет
        self.volatility_5y = expected_returns_and_volatility[1]  # Ожидаемая волатильность за 5 лет
        self.buffet_score = calculate_buffet_score.calculate_subjective_score(self.ticker)
        self.multipliers_score = calculate_multiplier_score.evaluate_company(self.ticker)["score"]  # Оценка через P/E, P/B, P/S
        self.vwap_score = calculate_vwap_score.evaluate_company(self.ticker)["score"]  # Оценка VWAP
        self.dividend_score = calculate_dividends_score.evaluate_company(self.ticker)["score"]  # Оценка стабильности дивидендов
        self.total_score = None  # Общая оценка на основе методов выше

        self.calculate_scores()        

    def calculate_scores(self):
        # Общая оценка акции
        self.total_score = sum(
            [
                self.buffet_score,
                self.multipliers_score,
                self.vwap_score,
                self.dividend_score
            ]
        )

    def get_moex_data(self):
        base_url = "https://iss.moex.com/iss"
        ticker = self.ticker

        try:
            # Доступ к текущей цене
            market_data_url = f"{base_url}/engines/stock/markets/shares/securities/{ticker}.json"
            market_data_response = requests.get(market_data_url)
            market_data = market_data_response.json()

            # Поиск элемента с рынком TQBR в данных
            data = market_data['securities']['data']
            tqbr_entry = next((entry for entry in data if entry[1] == "TQBR"), None)

            if tqbr_entry is None:
                raise ValueError("Не найден элемент с рынком TQBR")

            # Доступ к текущей цене (поле LAST на индексе 3)
            last_price = tqbr_entry[3]
            lot_size = tqbr_entry[4]
            # 2. Получение информации о компании (имя и сектор)
            security_data_url = f"{base_url}/securities/{ticker}.json"
            security_data_response = requests.get(security_data_url)
            security_data = security_data_response.json()

            # Имя компании
            company_name = security_data['description']['data'][1][2]  # "NAME" поле (индекс 2)
            
            self.current_price = last_price      
            self.company_name = company_name 
            self.lot_price = last_price * lot_size

            return {
                "ticker": ticker,
                "current_price": last_price,
                "company_name": company_name,
            }

        except Exception as e:
            return f"Ошибка получения данных для тикера {ticker}: {e}"

    def get_sector_name(self):
        url = f'https://smart-lab.ru/q/{self.ticker}/f/y/'
        response = requests.get(url)
        if response.status_code == 200:
            html = response.text

            # Парсинг HTML
            soup = BeautifulSoup(html, 'html.parser')
            # Поиск всех ссылок, где data-title начинается с "Aнализ сектора"
            links = soup.find_all('a', {'title': lambda value: value and value.startswith('Aнализ сектора')})
            return ''.join(c for c in links[0]["title"].split("сектор")[1] if c.isupper())
        else:
            return "Не определено"

class Sector:
    def __init__(self, name: str):
        self.name = name
        self.stocks: List[Stock] = []
        self.weights: List[float] = []  # Вес каждой акции в секторе
        self.calculate_average_score()

    def add_stock(self, stock: Stock):
        """Добавить акцию в сектор."""
        self.stocks.append(stock)
        self.calculate_weights()

    def remove_stock(self, ticker: str):
        """Удалить акцию из сектора по тикеру."""
        self.stocks = [stock for stock in self.stocks if stock.ticker != ticker]
        self.calculate_weights()

    def calculate_weights(self):
        """Рассчитать веса акций в секторе на основе их оценок."""
        total_score = sum(stock.total_score for stock in self.stocks)
        if total_score > 0:
            self.weights = [stock.total_score / total_score for stock in self.stocks]
        else:
            self.weights = [0] * len(self.stocks)
        self.calculate_average_score()

    def calculate_average_score(self):
        """Подсчитать среднюю оценку сектора."""
        if self.stocks:
            self.average_score = sum(stock.total_score for stock in self.stocks) / len(self.stocks)
        else:
            self.average_score = 0


class Portfolio:
    def __init__(self):
        self.sectors: List[Sector] = []
        self.sector_weights: List[float] = []  # Вес каждого сектора в портфеле
        self.stock_weights: Dict[str, float] = {}  # Вес каждой акции в портфеле
        self.real_stock_weights: Dict[str, float] = {}  # Вес каждой акции в портфеле

    def add_sector(self, sector: Sector):
        """Добавить сектор в портфель."""
        self.sectors.append(sector)
        self.calculate_sector_weights()

    def remove_stock(self, ticker: str):
        """Удалить акцию из портфеля по тикеру."""
        for sector in self.sectors:
            sector.remove_stock(ticker)
        self.calculate_sector_weights()

    def calculate_sector_weights(self):
        """Рассчитать веса секторов на основе их оценок."""
        total_score = sum(sector.average_score for sector in self.sectors)
        if total_score > 0:
            self.sector_weights = [sector.average_score / total_score for sector in self.sectors]
        else:
            self.sector_weights = [0] * len(self.sectors)
        self.calculate_stock_weights()

    def calculate_stock_weights(self):
        """Рассчитать веса акций в портфеле."""
        self.stock_weights = {}
        for sector, sector_weight in zip(self.sectors, self.sector_weights):
            for stock, stock_weight in zip(sector.stocks, sector.weights):
                self.stock_weights[stock.ticker] = sector_weight * stock_weight
        self.real_stock_weights = self.stock_weights
        self.calculate_min_portfolio_cost()

    def calculate_real_stock_weights(self, total_sum):
        """Рассчитать реальные веса акций в портфеле."""
        self.real_stock_weights = {}
        for sector in self.sectors:
            for stock in sector.stocks:
                self.real_stock_weights[stock.ticker] = stock.num_lots * stock.lot_price / total_sum

    def save_to_xlsx(self, filename: str):
        """Сохранить данные портфеля в Excel."""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Заголовки таблицы
        headers = [
            "Сектор", "Тикер", "Название компании", "Текущая цена", "ОД5Л",
            "В5Л", "Оценка Баффета", "Справедливая стоимость",
            "Мультипликаторы", "VWAP", "Дивиденды", "Общая оценка акции",
            "Оценка сектора", "Вес акции в портфеле", "Стоимость лота", "Кол-во лотов", "Реальные веса"
        ]
        ws.append(headers)

        for sector in self.sectors:
            for stock in sector.stocks:
                ws.append([
                    sector.name,
                    stock.ticker,
                    stock.company_name,
                    stock.current_price,
                    stock.return_5y,
                    stock.volatility_5y,
                    stock.buffet_score,
                    stock.multipliers_score,
                    stock.vwap_score,
                    stock.dividend_score,
                    stock.total_score,
                    sector.average_score,
                    self.stock_weights.get(stock.ticker, 0),
                    stock.lot_price,
                    stock.num_lots,
                    self.real_stock_weights.get(stock.ticker, 0)
                ])

        wb.save(filename)

    def load_from_xlsx(self, filename: str):
        """Загрузить данные портфеля из Excel."""
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        
        # Очищаем текущий портфель
        self.sectors = []

        # Читаем строки таблицы
        for row in ws.iter_rows(min_row=2, values_only=True):
            sector_name, ticker, company_name, current_price, *metrics = row
            
            # Создаем акцию
            stock = Stock(ticker)

            # Ищем или создаем сектор
            sector = next((s for s in self.sectors if s.name == sector_name), None)
            if not sector:
                sector = Sector(sector_name)
                self.add_sector(sector)

            sector.add_stock(stock)
            time.sleep(1)
        self.calculate_sector_weights()

    def calculate_min_portfolio_cost(self):
        # Нормализация весов, если их сумма не равна 1
        weights = np.array(list(self.stock_weights.values()))
        weights = weights / weights.sum()
        if weights.size != 0:
            prices = []
            for sector in self.sectors:
                for stock in sector.stocks:
                    prices.append(stock.lot_price)
            prices = np.array(prices)

            # Новый алгоритм
            quantities = weights * prices[np.argmax(prices)]/weights[np.argmax(prices)] / prices
            quantities[np.argmax(prices)] = 1

            quantities = np.ceil(quantities/quantities[np.argmin(quantities)])
            

            '''# Старый алгоритм
            quantities = weights * np.sum(prices/weights) / prices
            quantities = np.ceil(quantities).astype(int)
            quantities = np.ceil(quantities / np.gcd.reduce(quantities)).astype(int)'''

            # Шаг 3: Вычислить минимальную стоимость портфеля
            total_cost = np.sum(quantities * prices)

            i = 0
            for sector in self.sectors:
                for stock in sector.stocks:
                    stock.num_lots = quantities[i]
                    i += 1

            return total_cost

    def get_quantities_of_portfolio(self, total_sum):
        weights = np.array(list(self.stock_weights.values()))
        if weights.size != 0:
            prices = []
            for sector in self.sectors:
                for stock in sector.stocks:
                    prices.append(stock.lot_price)
            prices = np.array(prices)

            # 1. Расчет целевых капиталовложений для каждой акции
            target_investment = weights * total_sum

            # 2. Первоначальный расчет количества акций
            initial_quantities = target_investment / prices

            # 3. Округление количества акций до целых чисел (с учетом минимального отклонения)
            rounded_quantities = np.floor(initial_quantities).astype(int)  # Округление вниз для начала

            i = 0
            for sector in self.sectors:
                for stock in sector.stocks:
                    stock.num_lots = rounded_quantities[i]
                    i += 1

            # Вывод результатов
            return rounded_quantities